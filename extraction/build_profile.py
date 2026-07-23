"""Build a merged profile.json from official + community Profile.xlsx files.

The output JSON is the single source of truth consumed at runtime by
profile_resolver.  Custom overrides live separately in custom_profile.json.

Usage:
    python build_profile.py                          # uses default paths
    python build_profile.py official.xlsx [community.xlsx] [output.json]
"""

import json
import os
import sys
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ── XLSX Parsing ────────────────────────────────────────────────────────────

def parse_types_sheet(ws) -> Dict[str, Dict]:
    """Parse the 'Types' sheet into {type_name: {base_type, values}}."""
    types: Dict[str, Dict] = {}
    current_type: Optional[str] = None
    current_base: Optional[str] = None

    header = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(header) if h}

    def cell(row, name):
        idx = col.get(name)
        if idx is not None and idx < len(row):
            return row[idx].value
        return None

    for row in ws.iter_rows(min_row=2):
        type_name = cell(row, 'Type Name')
        base_type = cell(row, 'Base Type')
        value_name = cell(row, 'Value Name')
        value = cell(row, 'Value')
        comment = cell(row, 'Comment')

        if type_name is not None and str(type_name).strip():
            current_type = str(type_name).strip()
            current_base = str(base_type).strip() if base_type else None
            if current_type not in types:
                types[current_type] = {'base_type': current_base, 'values': {}}
            elif current_base:
                types[current_type]['base_type'] = current_base

        if current_type and value is not None and value_name is not None and str(value_name).strip():
            try:
                val_str = str(value).strip()
                int_val = int(val_str, 16) if val_str.startswith('0x') else int(float(val_str))
                entry: Dict[str, Any] = {'name': str(value_name).strip()}
                if comment and str(comment).strip():
                    entry['comment'] = str(comment).strip()
                types[current_type]['values'][int_val] = entry
            except (ValueError, TypeError):
                pass

    return types


def parse_messages_sheet(ws, types: Dict[str, Dict],
                        extra_mesg_map: Optional[Dict[str, int]] = None) -> Dict[int, Dict]:
    """Parse the 'Messages' sheet into {global_id: {name, fields}}.

    Args:
        extra_mesg_map: Optional pre-built mesg_name→id mapping (e.g. from the
            official profile) so that community messages referencing official
            message names can be resolved.
    """
    # Build mesg_name → id mapping from the mesg_num type
    mesg_name_to_id: Dict[str, int] = dict(extra_mesg_map or {})
    mesg_num_type = types.get('mesg_num', {})
    for val, entry in mesg_num_type.get('values', {}).items():
        mesg_name_to_id[entry['name']] = val

    header = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(header) if h}

    def cell(row, name):
        idx = col.get(name)
        if idx is not None and idx < len(row):
            return row[idx].value
        # Handle 'Products:' vs 'Products'
        if name == 'Products':
            return cell(row, 'Products:')
        return None

    messages: Dict[int, Dict] = {}
    current_msg_name: Optional[str] = None
    current_msg_id: Optional[int] = None
    last_field_def_num: Optional[int] = None

    for row in ws.iter_rows(min_row=2):
        msg_name = cell(row, 'Message Name')
        field_def_num = cell(row, 'Field Def #')
        field_name = cell(row, 'Field Name')
        field_type = cell(row, 'Field Type')

        # Also check 'Global Message Number' column (community xlsx)
        global_msg_num = cell(row, 'Global Message Number')

        if msg_name is not None and str(msg_name).strip():
            current_msg_name = str(msg_name).strip()
            current_msg_id = mesg_name_to_id.get(current_msg_name)
            # Fallback: use the 'Global Message Number' column if present
            if current_msg_id is None and global_msg_num is not None:
                try:
                    current_msg_id = int(float(str(global_msg_num).strip()))
                except (ValueError, TypeError):
                    pass
            last_field_def_num = None
            if current_msg_id is not None and current_msg_id not in messages:
                messages[current_msg_id] = {'name': current_msg_name, 'fields': {}}

        if current_msg_id is None or not (field_name and str(field_name).strip()):
            continue

        fi: Dict[str, Any] = {
            'name': str(field_name).strip(),
            'type': str(field_type).strip() if field_type and str(field_type).strip() else 'enum',
        }

        # Optional metadata columns
        opt_cols = [
            ('Scale', 'scale'), ('Offset', 'offset'), ('Units', 'units'),
            ('Array', 'array'), ('Components', 'components'), ('Bits', 'bits'),
            ('Accumulate', 'accumulate'), ('Ref Field Name', 'ref_field_name'),
            ('Ref Field Value', 'ref_field_value'), ('Comment', 'comment'),
            ('Products', 'products'), ('EXAMPLE', 'example'),
        ]
        for xlsx_col, key in opt_cols:
            val = cell(row, xlsx_col)
            if val is None:
                continue
            if isinstance(val, str):
                val = val.strip()
                if not val:
                    continue
            # Don't store trivial scale=1 / offset=0
            if key == 'scale' and val == 1:
                continue
            if key == 'offset' and val == 0:
                continue
            fi[key] = val

        if field_def_num is not None and str(field_def_num).strip():
            num = int(float(str(field_def_num).strip()))
            messages[current_msg_id]['fields'][num] = fi
            last_field_def_num = num
        else:
            # Subfield (no field_def_num → attach to parent)
            if 'ref_field_name' in fi and last_field_def_num is not None:
                parent = messages[current_msg_id]['fields'].get(last_field_def_num)
                if parent:
                    parent.setdefault('subfields', []).append(fi)

    return messages


def parse_profile_xlsx(path: str,
                       extra_mesg_map: Optional[Dict[str, int]] = None) -> Dict[str, Any]:
    """Parse a Profile.xlsx → {'types': …, 'messages': …}.

    Args:
        extra_mesg_map: Optional pre-built mesg_name→id mapping passed through
            to the messages parser (needed when parsing community xlsx so that
            message names defined only in the official mesg_num can still be
            resolved).
    """
    print(f"Parsing {path} …")
    wb = load_workbook(path, read_only=True, data_only=True)
    types = parse_types_sheet(wb['Types'])
    messages = parse_messages_sheet(wb['Messages'], types, extra_mesg_map)
    wb.close()
    t_count = len(types)
    tv_count = sum(len(t.get('values', {})) for t in types.values())
    m_count = len(messages)
    f_count = sum(len(m.get('fields', {})) for m in messages.values())
    print(f"  → {t_count} types ({tv_count} values), {m_count} messages ({f_count} fields)")
    return {'types': types, 'messages': messages}


# ── Merging ─────────────────────────────────────────────────────────────────

def merge_profiles(official: Dict, community: Dict) -> Dict:
    """Merge community additions into official profile (non-destructive)."""
    types = dict(official.get('types', {}))
    messages = dict(official.get('messages', {}))

    stats = {'new_types': 0, 'new_type_values': 0, 'new_messages': 0, 'new_fields': 0}

    for name, t in community.get('types', {}).items():
        if name.endswith('?'):
            continue
        if name not in types:
            types[name] = t
            stats['new_types'] += 1
            stats['new_type_values'] += len(t.get('values', {}))
        else:
            existing = types[name].setdefault('values', {})
            for k, v in t.get('values', {}).items():
                if v.get('name', '').endswith('?'):
                    continue
                if k not in existing:
                    existing[k] = v
                    stats['new_type_values'] += 1

    for msg_id, msg in community.get('messages', {}).items():
        if msg.get('name', '').endswith('?'):
            continue
        if msg_id not in messages:
            for fid, fdef in list(msg.get('fields', {}).items()):
                if fdef.get('name', '').endswith('?'):
                    del msg['fields'][fid]
            messages[msg_id] = msg
            stats['new_messages'] += 1
            stats['new_fields'] += len(msg.get('fields', {}))
        else:
            for fid, fdef in msg.get('fields', {}).items():
                if fdef.get('name', '').endswith('?'):
                    continue
                if fid not in messages[msg_id].get('fields', {}):
                    messages[msg_id].setdefault('fields', {})[fid] = fdef
                    stats['new_fields'] += 1

    print(f"Merge stats: +{stats['new_types']} types, +{stats['new_type_values']} values, "
          f"+{stats['new_messages']} messages, +{stats['new_fields']} fields")
    return {'types': types, 'messages': messages}


# ── JSON output ─────────────────────────────────────────────────────────────

def _to_json_dict(data: Dict) -> Dict:
    """Convert integer keys to strings for JSON serialisation."""
    types = {}
    for name, t in data['types'].items():
        tc = dict(t)
        if 'values' in tc:
            tc['values'] = {str(k): v for k, v in tc['values'].items()}
        types[name] = tc

    messages = {}
    for msg_id, msg in data['messages'].items():
        mc = dict(msg)
        if 'fields' in mc:
            mc['fields'] = {str(k): v for k, v in mc['fields'].items()}
        messages[str(msg_id)] = mc

    return {'types': types, 'messages': messages}


def build_profile(official_xlsx: str,
                  community_xlsx: Optional[str] = None,
                  output_path: Optional[str] = None) -> str:
    """Main entry point: parse, merge, write profile.json."""
    official = parse_profile_xlsx(official_xlsx)

    if community_xlsx and os.path.exists(community_xlsx):
        # Build official mesg_name→id mapping so the community parser can
        # resolve message names that only exist in the official mesg_num.
        official_mesg_map: Dict[str, int] = {}
        off_mesg_num = official.get('types', {}).get('mesg_num', {})
        for val, entry in off_mesg_num.get('values', {}).items():
            official_mesg_map[entry['name']] = val
        community = parse_profile_xlsx(community_xlsx, extra_mesg_map=official_mesg_map)
        merged = merge_profiles(official, community)
    else:
        merged = official

    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "profile.json"
        )

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(_to_json_dict(merged), f, indent=2)

    t = len(merged['types'])
    m = len(merged['messages'])
    print(f"Profile written: {t} types, {m} messages → {output_path}")
    return output_path


# ── CLI ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    argv = sys.argv[1:]
    off = argv[0] if len(argv) > 0 else os.path.join(base, 'Profile_Official.xlsx')
    com = argv[1] if len(argv) > 1 else os.path.join(base, 'Profile_Community.xlsx')
    out = argv[2] if len(argv) > 2 else None
    build_profile(off, com, out)
